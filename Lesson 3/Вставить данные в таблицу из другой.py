from openpyxl import load_workbook                                              # Импортировали из библиотеки метод для загрузки файлов Excel
from openpyxl.chart import (AreaChart, Reference)                               # Импортировали некоторые нужные классы для работы с графиками

wb1 = load_workbook(filename = 'test.xlsx')                                     # Загружаем книгу test.xlsx
wb2 = load_workbook(filename = 'test2.xlsx')                                    # Загружаем книгу test2.xlsx

sheet = wb1['Sheet']                                                            # В переменную sheet записываем ссылку на лист Sheet
r = sheet_ranges["A1:C2"]                                                       # В переменную r записали массив значений из диапазона A1:C2

l = []                                                                          # Массив для хранения значений из ячеек диапазона r
for i in range(len(r)):                                                         # Для каждого i из диапазона от 0 до количества элементов в r
    l.append([])                                                                # Создаём пустой список
    for j in r[i]:                                                              # Для каждого j из элементов r[i]
        l[i].append(j.value)                                                    # Добавляем в конец массива l[i] значение j из r[i]
    

ws2 = wb2.active                                                                # Создаём переменную для первого листа книги wb2
for i in l:                                                                     # Для каждого i из списка l
    ws2.append(i)                                                               # Добавляем строку i в лист ws2

chart = AreaChart()                                                             # Создаём объект графика
chart.title = "Area Chart"                                                      # Задаём ему заголовок
chart.style = 13                                                                # 
chart.x_axis.title = 'Test'                                                     # Задаём заголовок оси x
chart.y_axis.title = 'Percentage'                                               # Задаём заголовок оси y

cats = Reference(ws2, min_col=1, min_row=1, max_row=ws2.max_row)                # Задаём значения для оси x
data = Reference(ws2, min_col=2, min_row=1, max_col=3, max_row=ws2.max_row)     # Задаём данные для графика
chart.add_data(data)                                                            # Добавляем данные в график
chart.set_categories(cats)                                                      # Добавляем графику подписи

ws2.add_chart(chart, "A10")                                                     # Добавляем график в ячейку A10

wb2.save('test2.xlsx')                                                          # Сохраняем файл