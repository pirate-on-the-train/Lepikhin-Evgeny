# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('данные на 01.08.2020.xlsx')

# Get sheet names
print(wb.get_sheet_names())