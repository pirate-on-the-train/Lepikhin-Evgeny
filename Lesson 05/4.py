
from openpyxl import Workbook, load_workbook

wb1 = load_workbook(filename = '1.xlsx')
wb2 = load_workbook(filename = '2.xlsx')

ws1 = wb1.active
ws2 = wb2.active

heads = ws2['A']

print([i.value for i in heads])

#for i in heads:
    #a = lambda m: m.value
    #if '3' in [for j in heads a]:
    #    print(1)