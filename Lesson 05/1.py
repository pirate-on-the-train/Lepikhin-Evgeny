'''
ТЗ что нужно сделать:
1. добавить в базу данных показатели по скважинам на 01.08.2020;
2. провести суммирование показателей Дебит жидкости (ТР), м3/сут, Дебит нефти (ТР), т/сут и Дебит попутного газа, тыс. м3/сут на каждую дату и построить график по всему месторождению;
3. найти по месторождению даты максимальной добычи нефти и попутного газа, вероятно они не будут совпадать, но не факт;
4. провести усреднение отдельно по нефтяным и нагнетальным скважинам (столбец характер работы) на каждую дату пластовое давление, забойное давление и устьевое давление и построить 2 графика по этим данным, т.е. отдельно по нагнетательным...отдельно по нефтяным;
5. провести расчет количества скважин проперфорированных на Ю1-3(столбец перфорированые пласты) по дате и построить гистограмму.
'''

from openpyxl import Workbook, load_workbook

#wb1 = load_workbook(filename = 'база данных.xlsx')
wb2 = load_workbook(filename = 'данные на 01.08.2020.xlsx')

#ws1 = wb1.active
ws2 = wb2.active


ss = {
    'A' : 1,
    'B' : 2,
    'C' : 3
}

l = []
for col in range(1, ws2.max_column + 1):
    # l.append(ws2.cell(column=col, row=1).value)
    column = ws2.cell(column=col, row=1).value
    ws1.

for i in l:
    for j in range(1, ws1.max_column + 1):
        if 
print(l)