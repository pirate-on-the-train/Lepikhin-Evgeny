from openpyxl import Workbook                   # Импортируем класс Workbook из модуля openpyxl
from openpyxl.utils import get_column_letter    # Импортируем метод get_column_letter из модуля openpyxl.utils

wb = Workbook()                                 # Создаём переменную wb для работы с книгой

ws = wb.active                                  # Создаём переменную ws, в которую записываем активный лист (т.е. первую)
ws['A4'] = 7                                    # В ячейку A4 листа ws записываем 7

ws2 = wb.create_sheet('22')                     # Создаём второй лист в книге wb и записываем его в ws2

rows = [                                        # Список rows состоящий из 2 других списков 
    [1, 2, 3],                                  # 1 строка таблицы
    [4, 5, 6]                                   # 2 строка таблицы
]

for row in rows:                                # Для каждого элемента row в списке rows
    ws2.append(row)                             # На лист ws2 добавили каждую строку row

sheet_ranges = wb['Sheet']                      # Получаем лист с названием 'Sheet' и записываем в переменную sheet_ranges
ws2.append(list(sheet_ranges['A4'].value))      # Добавляем в лист ws2 строку со значением из ячейки A4 листа sheet_ranges (первый лист Sheet)

wb.save('test.xlsx')                            # Сохраняем книгу