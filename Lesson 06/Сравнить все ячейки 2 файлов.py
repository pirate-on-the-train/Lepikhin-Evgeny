from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter                                            # Подключили функцию get_column_letter из модуля openpyxl.utils

wb1 = load_workbook(filename = '1.xlsx')
wb2 = load_workbook(filename = '2.xlsx')

ws1 = wb1.active
ws2 = wb2.active


for row1 in range(1, ws1.max_row + 1):
    for col1 in range(1, ws1.max_column + 1):
        a = ws1.cell(column = col1, row = row1)
        for row2 in range(1, ws2.max_row + 1):
            for col2 in range(1, ws2.max_column + 1):
                b = ws2.cell(column = col2, row = row2)
                if a.value == b.value:
                    print('Есть сходство в ячейке ({0},{1})'.format(b.row, b.column))