'''
ТЗ что нужно сделать:
1. добавить в базу данных показатели по скважинам на 01.08.2020;
2. провести суммирование показателей Дебит жидкости (ТР), м3/сут, Дебит нефти (ТР), т/сут и Дебит попутного газа, тыс. м3/сут на каждую дату и построить график по всему месторождению;
3. найти по месторождению даты максимальной добычи нефти и попутного газа, вероятно они не будут совпадать, но не факт;
4. провести усреднение отдельно по нефтяным и нагнетальным скважинам (столбец характер работы) на каждую дату пластовое давление, забойное давление и устьевое давление и построить 2 графика по этим данным, т.е. отдельно по нагнетательным...отдельно по нефтяным;
5. провести расчет количества скважин проперфорированных на Ю1-3(столбец перфорированые пласты) по дате и построить гистограмму.
'''

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter                                            # Подключили функцию get_column_letter из модуля openpyxl.utils

wb1 = load_workbook(filename = '1.xlsx')
wb2 = load_workbook(filename = '2.xlsx')

ws1 = wb1.active
ws2 = wb2.active

heads = ws1[1]
dates = ws1['A']

bb = {}
for i in heads:
    bb[i.value] = i.column_letter


for row in range(2, ws2.max_row + 1):
    mr = ws1.max_row
    v = ws2[row]
    
    qasd = [i.value for i in v]
    #qasd1 = [[i.value for i in ws1[i.column_letter]] for j in heads]
    #qasd1 = [[ws1.cell(column=col, row=row) for col in ws1[i.column_letter]] for row in heads]
    #zxcvb = [[ws2[i][j] for i in range(2, ws2.max_row + 1)].value for j in range(ws2.max_column)]
    
    qasd1 = [[ws1[j][i].value for i in range(ws1.max_column)] for j in range(2, ws1.max_row + 1)]
    
    if qasd in qasd1:
        print('Yes')
    
    #if [i.value for i in v]    v[col].value in [i.value for i in ws1[get_column_letter(col)]]:
        #if v[col].value in [i.value for i in ws1[get_column_letter(col)]]:
            #print('такой элемент уже есть')
    
    
    
    for col in range(0, ws2.max_column):
        zz = v[col].value
        
        for col in range(1, ws2.max_column + 1):
            column = ws2.cell(column=col, row=1).value
            for i in heads:
                if i.value == column:
                    value = ws2.cell(column=col, row=row).value                
                    ws1.cell(column=heads.index(i) + 1, row = mr + 1, value = value)
                    break

wb1.save('3.xlsx')